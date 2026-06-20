#!/usr/bin/env python3
"""
Suggest and apply Excel region updates from public/images/places/ folder structure.

Only updates when folder name match confidence is high (exact, alias, or single normalized).
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install openpyxl: python3 -m pip install openpyxl", file=sys.stderr)
    sys.exit(1)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_EXCEL = PROJECT_ROOT / "data" / "imports" / "places-filter.xlsx"
PLACES_ROOT = PROJECT_ROOT / "public" / "images" / "places"
REPORT_FILE = PROJECT_ROOT / "data" / "excel-region-folder-report.json"

FOLDER_TO_REGION = {
    "north": "צפון",
    "hasharon": "השרון",
    "center": "מרכז",
    "jerusalem": "ירושלים",
    "south": "דרום",
}

PLACE_COLUMN = "שם המקום"
REGION_COLUMN = "אזור"

EXCEL_TO_FOLDER_ALIASES: dict[str, str] = {
    "נחל שופט": "נחל השופט",
    "שמורת תל דן": "נחל דן",
    "פארק המעיינות": "עין מודע",
    "שמורת טבע נחל שורק": "שפך נחל שורק",
    "עין חרוד": "מעיין חרוד",
    "פארק נחל לכיש": "פארק לכיש",
}

FOLDER_TO_EXCEL_ALIASES: dict[str, str] = {
    "פארק נחל לכיש": "פארק לכיש",
}


def normalize_title(value: str) -> str:
    text = value.strip()
    text = re.sub(r"[\s\-–—_/\\|]", "", text)
    text = re.sub(r"^ה", "", text)
    return text.casefold()


def scan_image_folders() -> dict[str, str]:
    folder_regions: dict[str, str] = {}
    for slug, region_title in FOLDER_TO_REGION.items():
        region_dir = PLACES_ROOT / slug
        if not region_dir.is_dir():
            continue
        for entry in region_dir.iterdir():
            if entry.is_dir():
                folder_regions[entry.name] = region_title
    return folder_regions


def resolve_folder_match(
    excel_name: str,
    folder_regions: dict[str, str],
) -> tuple[str | None, str | None, str | None]:
    """Return (folder_name, folder_region, match_kind) or (None, None, None)."""

    candidates: list[tuple[str, str, str]] = []

    if excel_name in folder_regions:
        candidates.append((excel_name, folder_regions[excel_name], "exact"))

    alias_folder = EXCEL_TO_FOLDER_ALIASES.get(excel_name)
    if alias_folder and alias_folder in folder_regions:
        candidates.append((alias_folder, folder_regions[alias_folder], "excel-alias"))

    folder_alias = FOLDER_TO_EXCEL_ALIASES.get(excel_name)
    if folder_alias and folder_alias in folder_regions:
        candidates.append((folder_alias, folder_regions[folder_alias], "folder-alias"))

    normalized_hits = [
        folder_name
        for folder_name in folder_regions
        if normalize_title(folder_name) == normalize_title(excel_name)
    ]
    if len(normalized_hits) == 1:
        folder_name = normalized_hits[0]
        candidates.append((folder_name, folder_regions[folder_name], "normalized"))

    if not candidates:
        partial_hits = [
            folder_name
            for folder_name in folder_regions
            if normalize_title(excel_name) in normalize_title(folder_name)
            or normalize_title(folder_name) in normalize_title(excel_name)
        ]
        if len(partial_hits) == 1:
            folder_name = partial_hits[0]
            return folder_name, folder_regions[folder_name], "partial-review"
        if len(partial_hits) > 1:
            return None, None, "multiple-partial"
        return None, None, None

    unique_folders = {item[0] for item in candidates}
    unique_regions = {item[1] for item in candidates}
    if len(unique_folders) > 1 or len(unique_regions) > 1:
        return None, None, "conflicting-folder-matches"

    return candidates[0]


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Update Excel region column from image folder structure"
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    excel_path = args.input
    if not excel_path.is_file():
        print(f"Excel file not found: {excel_path}", file=sys.stderr)
        return 1

    folder_regions = scan_image_folders()
    workbook = load_workbook(excel_path)
    sheet = workbook.active

    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    if PLACE_COLUMN not in headers or REGION_COLUMN not in headers:
        print(f"Missing required columns. Found: {headers}", file=sys.stderr)
        return 1

    place_idx = headers.index(PLACE_COLUMN)
    region_idx = headers.index(REGION_COLUMN)

    updated: list[dict] = []
    unchanged: list[str] = []
    manual_review: list[dict] = []
    matched_folders: set[str] = set()

    for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        place_cell = row[place_idx] if place_idx < len(row) else None
        region_cell = row[region_idx] if region_idx < len(row) else None
        if place_cell is None or place_cell.value is None:
            continue

        excel_name = str(place_cell.value).strip()
        excel_region = str(region_cell.value).strip() if region_cell and region_cell.value else ""
        if not excel_name:
            continue

        folder_name, folder_region, match_kind = resolve_folder_match(excel_name, folder_regions)

        if match_kind in {"partial-review", "multiple-partial", "conflicting-folder-matches"}:
            manual_review.append(
                {
                    "excelName": excel_name,
                    "excelRegion": excel_region,
                    "reason": match_kind,
                    "folderName": folder_name,
                    "folderRegion": folder_region,
                }
            )
            unchanged.append(excel_name)
            continue

        if folder_name is None or folder_region is None:
            unchanged.append(excel_name)
            continue

        matched_folders.add(folder_name)

        if folder_region == excel_region:
            unchanged.append(excel_name)
            continue

        if match_kind in {"exact", "excel-alias", "folder-alias", "normalized"}:
            if not args.dry_run and region_cell is not None:
                region_cell.value = folder_region
            updated.append(
                {
                    "excelName": excel_name,
                    "fromRegion": excel_region,
                    "toRegion": folder_region,
                    "folderName": folder_name,
                    "matchKind": match_kind,
                    "row": row_number,
                }
            )
        else:
            manual_review.append(
                {
                    "excelName": excel_name,
                    "excelRegion": excel_region,
                    "reason": match_kind or "unknown",
                    "folderName": folder_name,
                    "folderRegion": folder_region,
                }
            )

    folders_without_excel = [
        {
            "folderName": folder_name,
            "folderRegion": folder_regions[folder_name],
        }
        for folder_name in sorted(folder_regions)
        if folder_name not in matched_folders
    ]

    report = {
        "excelFile": str(excel_path),
        "updatedCount": len(updated),
        "unchangedCount": len(unchanged),
        "manualReviewCount": len(manual_review),
        "foldersWithoutExcelCount": len(folders_without_excel),
        "updated": updated,
        "manualReview": manual_review,
        "foldersWithoutExcel": folders_without_excel,
    }

    if not args.dry_run and updated:
        workbook.save(excel_path)

    REPORT_FILE.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Excel file: {excel_path.resolve()}")
    print(f"Report: {REPORT_FILE}")
    print(f"Mode: {'dry-run' if args.dry_run else 'write'}")
    print("")
    print(f"Rows updated by folder region: {len(updated)}")
    for entry in updated:
        print(
            f"  - {entry['excelName']}: {entry['fromRegion']} → {entry['toRegion']} "
            f"({entry['matchKind']}, folder: {entry['folderName']})"
        )
    print("")
    print(f"Rows unchanged: {len(unchanged)}")
    print(f"Rows needing manual review: {len(manual_review)}")
    for entry in manual_review:
        print(f"  - {entry['excelName']!r}: {entry['reason']}")
    print("")
    print(f"Image folders without Excel row: {len(folders_without_excel)}")
    for entry in folders_without_excel:
        print(f"  - {entry['folderName']!r} ({entry['folderRegion']})")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
