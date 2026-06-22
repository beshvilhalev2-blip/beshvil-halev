#!/usr/bin/env python3
"""
Import visited places from data/imports/visited-places.xlsx into trip data.

- Matches existing rawTrips in data/trips.ts by title + region (no duplicates)
- Generates lightweight needs-content trips for unmatched places
- Writes data/visited-place-trips.ts and data/visited-place-matches.ts
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install openpyxl: python3 -m pip install openpyxl", file=sys.stderr)
    sys.exit(1)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_EXCEL_FILE = PROJECT_ROOT / "data" / "imports" / "visited-places.xlsx"
TRIPS_FILE = PROJECT_ROOT / "data" / "trips.ts"
OUTPUT_TRIPS_FILE = PROJECT_ROOT / "data" / "visited-place-trips.ts"
OUTPUT_MATCHES_FILE = PROJECT_ROOT / "data" / "visited-place-matches.ts"
PLACE_COLUMN = "שם המקום"
REGION_COLUMN = "אזור"

REGION_HERO_BACKGROUNDS: dict[str, str] = {
    "צפון": """
  linear-gradient(160deg, rgba(6, 78, 59, 0.75) 0%, rgba(15, 118, 110, 0.55) 50%, rgba(28, 25, 23, 0.7) 100%),
  url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='1920' height='480' viewBox='0 0 1920 480'%3E%3Cdefs%3E%3ClinearGradient id='g' x1='0%25' y1='0%25' x2='100%25' y2='100%25'%3E%3Cstop offset='0%25' stop-color='%23059669'/%3E%3Cstop offset='100%25' stop-color='%23134e4a'/%3E%3C/linearGradient%3E%3C/defs%3E%3Crect fill='url(%23g)' width='1920' height='480'/%3E%3C/svg%3E")
""".strip(),
    "מרכז": """
  linear-gradient(160deg, rgba(180, 83, 9, 0.75) 0%, rgba(217, 119, 6, 0.55) 50%, rgba(28, 25, 23, 0.7) 100%),
  url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='1920' height='480' viewBox='0 0 1920 480'%3E%3Cdefs%3E%3ClinearGradient id='g' x1='0%25' y1='0%25' x2='100%25' y2='100%25'%3E%3Cstop offset='0%25' stop-color='%23f59e0b'/%3E%3Cstop offset='100%25' stop-color='%23b45309'/%3E%3C/linearGradient%3E%3C/defs%3E%3Crect fill='url(%23g)' width='1920' height='480'/%3E%3C/svg%3E")
""".strip(),
    "ירושלים": """
  linear-gradient(160deg, rgba(87, 83, 78, 0.8) 0%, rgba(180, 83, 9, 0.5) 50%, rgba(28, 25, 23, 0.7) 100%),
  url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='1920' height='480' viewBox='0 0 1920 480'%3E%3Cdefs%3E%3ClinearGradient id='g' x1='0%25' y1='0%25' x2='100%25' y2='100%25'%3E%3Cstop offset='0%25' stop-color='%23a8a29e'/%3E%3Cstop offset='100%25' stop-color='%2344403c'/%3E%3C/linearGradient%3E%3C/defs%3E%3Crect fill='url(%23g)' width='1920' height='480'/%3E%3C/svg%3E")
""".strip(),
    "דרום": """
  linear-gradient(160deg, rgba(154, 52, 18, 0.8) 0%, rgba(194, 65, 12, 0.55) 50%, rgba(28, 25, 23, 0.7) 100%),
  url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='1920' height='480' viewBox='0 0 1920 480'%3E%3Cdefs%3E%3ClinearGradient id='g' x1='0%25' y1='0%25' x2='100%25' y2='100%25'%3E%3Cstop offset='0%25' stop-color='%23f97316'/%3E%3Cstop offset='100%25' stop-color='%237c2d12'/%3E%3C/linearGradient%3E%3C/defs%3E%3Crect fill='url(%23g)' width='1920' height='480'/%3E%3C/svg%3E")
""".strip(),
}

TRANSLIT = {
    "א": "",
    "ב": "b",
    "ג": "g",
    "ד": "d",
    "ה": "h",
    "ו": "v",
    "ז": "z",
    "ח": "ch",
    "ט": "t",
    "י": "y",
    "כ": "k",
    "ך": "k",
    "ל": "l",
    "מ": "m",
    "ם": "m",
    "נ": "n",
    "ן": "n",
    "ס": "s",
    "ע": "",
    "פ": "p",
    "ף": "f",
    "צ": "tz",
    "ץ": "tz",
    "ק": "k",
    "ר": "r",
    "ש": "sh",
    "ת": "t",
    "'": "",
    "׳": "",
    "״": "",
    '"': "",
}


def find_excel_file(explicit: Path | None) -> Path:
    if explicit:
        if not explicit.is_file():
            raise FileNotFoundError(explicit)
        return explicit

    if not DEFAULT_EXCEL_FILE.is_file():
        raise FileNotFoundError(
            f"Could not find project Excel file: {DEFAULT_EXCEL_FILE}"
        )

    return DEFAULT_EXCEL_FILE


def load_published_trips(trips_file: Path) -> list[dict[str, str]]:
    text = trips_file.read_text(encoding="utf-8")
    start = text.index("const rawTrips")
    end = text.index("export const trips")
    chunk = text[start:end]

    trips: list[dict[str, str]] = []
    pattern = re.compile(
        r'slug:\s*"([^"]+)"\s*,\s*\n\s*title:\s*"([^"]+)"\s*,\s*\n\s*subtitle:[\s\S]*?\n\s*region:\s*"([^"]+)"'
    )
    for match in pattern.finditer(chunk):
        trips.append(
            {
                "slug": match.group(1),
                "title": match.group(2),
                "region": match.group(3),
            }
        )
    return trips


def load_existing_lightweight_slugs(output_file: Path) -> dict[tuple[str, str], str]:
    if not output_file.is_file():
        return {}

    text = output_file.read_text(encoding="utf-8")
    mapping: dict[tuple[str, str], str] = {}
    blocks = re.findall(
        r'title:\s*"([^"]+)"\s*,\s*\n\s*slug:\s*"([^"]+)"\s*,\s*\n\s*region:\s*"([^"]+)"',
        text,
    )
    for title, slug, region in blocks:
        mapping[(title, region)] = slug
    return mapping


def transliterate_hebrew(text: str) -> str:
    result: list[str] = []
    for char in text:
        lower = char.lower()
        if lower in TRANSLIT:
            result.append(TRANSLIT[lower])
        elif char.isascii() and (char.isalnum() or char in "- "):
            result.append(char.lower())
        elif char in " -–-_/\\|":
            result.append("-")
    slug = "".join(result)
    slug = re.sub(r"[^a-z0-9-]+", "-", slug)
    slug = re.sub(r"-{2,}", "-", slug).strip("-")
    return slug or "place"


def make_visited_slug(
    title: str,
    region: str,
    used: set[str],
    existing: dict[tuple[str, str], str],
) -> str:
    key = (title, region)
    if key in existing:
        slug = existing[key]
        used.add(slug)
        return slug

    base = transliterate_hebrew(title) or transliterate_hebrew(region) + "-place"
    slug = base
    suffix = 2
    while slug in used:
        slug = f"{base}-{suffix}"
        suffix += 1
    used.add(slug)
    return slug


def read_places(excel_path: Path) -> list[tuple[str, str]]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        raise ValueError("Excel file is empty.")

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    place_idx = headers.index(PLACE_COLUMN)
    region_idx = headers.index(REGION_COLUMN)

    places: list[tuple[str, str]] = []
    seen: set[tuple[str, str]] = set()

    for row in rows[1:]:
        if not row:
            continue
        place_raw = row[place_idx] if place_idx < len(row) else None
        region_raw = row[region_idx] if region_idx < len(row) else None
        if place_raw is None or region_raw is None:
            continue

        title = str(place_raw).strip()
        region = str(region_raw).strip()
        if not title or not region:
            continue

        key = (title, region)
        if key in seen:
            continue
        seen.add(key)
        places.append((title, region))

    return places


def escape_ts(value: str) -> str:
    return value.replace("\\", "\\\\").replace('"', '\\"')


def render_lightweight_trip(entry: dict) -> str:
    title = escape_ts(entry["title"])
    slug = escape_ts(entry["slug"])
    region = escape_ts(entry["region"])
    hero = entry["heroBackgroundImage"]
    meta = escape_ts(
        f"{entry['title']} - מילאנה והילדים טיילו כאן. כתבה מלאה, תמונות וטיפים יעלו בהמשך."
    )

    return f"""  {{
    slug: "{slug}",
    title: "{title}",
    subtitle: "מילאנה והילדים טיילו כאן - כתבה מלאה בקרוב",
    region: "{region}",
    category: "טיול משפחתי",
    status: "needs-content",
    visitedByMilana: true,
    metaDescription: "{meta}",
    heroImageLabel: "תמונת רקע - {title}",
    heroBackgroundImage: `
{hero}
    `,
    about: [],
    personalStory: [],
    cost: [],
    tips: [],
    gallery: [],
    gallerySubtitle: "",
    nearbyPlaces: [],
    nearbySubtitle: "",
  }},"""


def render_lightweight_trips_file(entries: list[dict]) -> str:
    body = "\n".join(render_lightweight_trip(entry) for entry in entries)
    return f"""/**
 * Lightweight trip entries for visited places without full articles yet.
 *
 * Generated by scripts/import-visited-places.py - re-run after Excel updates.
 * Do not edit manually; changes will be overwritten on re-import.
 */

export const visitedPlaceTrips = [
{body}
] as const;
"""


def render_matches_file(matched_slugs: list[str]) -> str:
    lines = "\n".join(f'  "{escape_ts(slug)}",' for slug in matched_slugs)
    return f"""/**
 * Existing published trips matched from the visited-places Excel import.
 *
 * Generated by scripts/import-visited-places.py.
 */

export const visitedByMilanaMatchedSlugs = [
{lines}
] as const;

export type VisitedByMilanaMatchedSlug =
  (typeof visitedByMilanaMatchedSlugs)[number];
"""


def build_import_result(
    places: list[tuple[str, str]],
    published_trips: list[dict[str, str]],
    existing_slugs: dict[tuple[str, str], str],
) -> tuple[list[str], list[dict]]:
    by_title_region = {
        (trip["title"], trip["region"]): trip for trip in published_trips
    }
    used_slugs = {trip["slug"] for trip in published_trips}
    matched_slugs: list[str] = []
    lightweight_entries: list[dict] = []

    for title, region in places:
        published = by_title_region.get((title, region))
        if published:
            matched_slugs.append(published["slug"])
            continue

        slug = make_visited_slug(title, region, used_slugs, existing_slugs)
        hero = REGION_HERO_BACKGROUNDS.get(
            region, REGION_HERO_BACKGROUNDS["מרכז"]
        )
        lightweight_entries.append(
            {
                "title": title,
                "slug": slug,
                "region": region,
                "heroBackgroundImage": hero,
            }
        )

    return matched_slugs, lightweight_entries


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Import Excel visited places into trip data files"
    )
    parser.add_argument("--input", type=Path, help="Path to Excel file")
    args = parser.parse_args()

    excel_path = find_excel_file(args.input)
    published_trips = load_published_trips(TRIPS_FILE)
    existing_slugs = load_existing_lightweight_slugs(OUTPUT_TRIPS_FILE)

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(values_only=True), ())
    workbook.close()
    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]

    for column in (PLACE_COLUMN, REGION_COLUMN):
        if column not in headers:
            print(f"Missing required column: {column}", file=sys.stderr)
            print(f"Columns found: {headers}", file=sys.stderr)
            return 1

    places = read_places(excel_path)
    if not places:
        print("No places found.", file=sys.stderr)
        return 1

    matched_slugs, lightweight_entries = build_import_result(
        places, published_trips, existing_slugs
    )

    OUTPUT_TRIPS_FILE.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_TRIPS_FILE.write_text(
        render_lightweight_trips_file(lightweight_entries),
        encoding="utf-8",
    )
    OUTPUT_MATCHES_FILE.write_text(
        render_matches_file(matched_slugs),
        encoding="utf-8",
    )

    regions = sorted({region for _, region in places})
    region_counts = {
        region: sum(1 for _, r in places if r == region) for region in regions
    }

    print(f"Excel file: {excel_path.resolve()}")
    print(f"Columns found: {PLACE_COLUMN}, {REGION_COLUMN}")
    print(f"Total Excel places: {len(places)}")
    print(f"Matched existing trips: {len(matched_slugs)}")
    print(f"New lightweight trips: {len(lightweight_entries)}")
    print(f"Duplicates avoided: {len(matched_slugs)}")
    print(f"Wrote: {OUTPUT_TRIPS_FILE}")
    print(f"Wrote: {OUTPUT_MATCHES_FILE}")
    for region in regions:
        print(f"  {region}: {region_counts[region]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
