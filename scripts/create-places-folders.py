#!/usr/bin/env python3
"""
Create Desktop/מקומות/{region}/{place}/ folders from רשימת מקומות.xlsx.

Columns expected: שם המקום, אזור
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print(
        "Missing dependency: openpyxl\n"
        "Install with: python3 -m pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(1)

EXCEL_FILENAME = "רשימת מקומות.xlsx"
DEST_FOLDER_NAME = "מקומות"
PLACE_COLUMN = "שם המקום"
REGION_COLUMN = "אזור"

KNOWN_REGIONS = ("צפון", "מרכז", "ירושלים", "דרום", "שטח 4x4")
INVALID_PATH_CHARS = re.compile(r'[/\\:\0]')


def sanitize_folder_name(name: str) -> str:
    cleaned = INVALID_PATH_CHARS.sub("", name.strip())
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned


def find_excel_file(explicit_path: Path | None) -> Path:
    if explicit_path:
        if not explicit_path.is_file():
            raise FileNotFoundError(f"Excel file not found: {explicit_path}")
        return explicit_path

    search_roots = [
        Path.cwd(),
        Path.home() / "Desktop",
        Path.home() / "Downloads",
        Path(__file__).resolve().parent,
        Path(__file__).resolve().parent.parent,
    ]

    checked: list[Path] = []
    for root in search_roots:
        candidate = root / EXCEL_FILENAME
        checked.append(candidate)
        if candidate.is_file():
            return candidate

    checked_lines = "\n".join(f"  - {path}" for path in checked)
    raise FileNotFoundError(
        f"Could not find {EXCEL_FILENAME}. Searched:\n{checked_lines}\n"
        f"Pass the file path explicitly: --input /path/to/{EXCEL_FILENAME}"
    )


def read_places(excel_path: Path) -> dict[str, list[str]]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook.active

    rows = sheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        raise ValueError("Excel file is empty.")

    headers = [
        str(cell).strip() if cell is not None else "" for cell in header_row
    ]

    try:
        place_idx = headers.index(PLACE_COLUMN)
        region_idx = headers.index(REGION_COLUMN)
    except ValueError as error:
        raise ValueError(
            f"Expected columns '{PLACE_COLUMN}' and '{REGION_COLUMN}'. "
            f"Found: {headers}"
        ) from error

    places_by_region: dict[str, list[str]] = defaultdict(list)
    seen: set[tuple[str, str]] = set()

    for row in rows:
        if not row:
            continue

        place_raw = row[place_idx] if place_idx < len(row) else None
        region_raw = row[region_idx] if region_idx < len(row) else None

        if place_raw is None or region_raw is None:
            continue

        place = sanitize_folder_name(str(place_raw))
        region = sanitize_folder_name(str(region_raw))

        if not place or not region:
            continue

        key = (region, place)
        if key in seen:
            continue

        seen.add(key)
        places_by_region[region].append(place)

    workbook.close()
    return dict(places_by_region)


def create_folders(
    places_by_region: dict[str, list[str]], destination_root: Path
) -> tuple[int, int]:
    destination_root.mkdir(parents=True, exist_ok=True)

    regions_created = 0
    places_created = 0

    ordered_regions = [
        region for region in KNOWN_REGIONS if region in places_by_region
    ]
    extra_regions = sorted(
        region for region in places_by_region if region not in KNOWN_REGIONS
    )
    all_regions = ordered_regions + extra_regions

    for region in all_regions:
        region_path = destination_root / region
        region_path.mkdir(exist_ok=True)
        regions_created += 1

        for place in places_by_region[region]:
            (region_path / place).mkdir(exist_ok=True)
            places_created += 1

    return regions_created, places_created


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Create Desktop/מקומות folders from רשימת מקומות.xlsx"
    )
    parser.add_argument(
        "--input",
        type=Path,
        help=f"Path to {EXCEL_FILENAME} (auto-detected if omitted)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path.home() / "Desktop" / DEST_FOLDER_NAME,
        help=f"Destination root (default: ~/Desktop/{DEST_FOLDER_NAME})",
    )
    args = parser.parse_args()

    excel_path = find_excel_file(args.input)
    places_by_region = read_places(excel_path)

    if not places_by_region:
        print("No places found in the Excel file.", file=sys.stderr)
        return 1

    regions_count, places_count = create_folders(places_by_region, args.output)

    print("Done.")
    print(f"Regions: {regions_count}")
    print(f"Places: {places_count}")
    print(f"Destination: {args.output}")
    print(f"Source: {excel_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
