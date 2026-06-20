#!/usr/bin/env python3
"""
Sync trip regions and filter tags from data/imports/places-filter.xlsx.

Generates data/places-filter-sync.ts and prints an audit report.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install openpyxl: python3 -m pip install openpyxl", file=sys.stderr)
    sys.exit(1)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_EXCEL = PROJECT_ROOT / "data" / "imports" / "places-filter.xlsx"
TRIPS_FILE = PROJECT_ROOT / "data" / "trips.ts"
VISITED_TRIPS_FILE = PROJECT_ROOT / "data" / "visited-place-trips.ts"
OUTPUT_FILE = PROJECT_ROOT / "data" / "places-filter-sync.ts"

COLUMNS = {
    "שם המקום": "name",
    "אזור": "region",
    "מים": "water",
    "תצפית": "viewpoint",
    "קאמפינג": "camping",
    "4x4": "offroad",
    "נגיש לעגלות": "strollerAccessible",
    "חינם": "free",
}

REGION_TO_SLUG = {
    "צפון": "north",
    "השרון": "hasharon",
    "מרכז": "center",
    "ירושלים": "jerusalem",
    "דרום": "south",
    "יהודה ושומרון": "judea-samaria",
}

REGION_TO_TITLE = {
    "north": "צפון",
    "hasharon": "השרון",
    "center": "מרכז",
    "jerusalem": "ירושלים",
    "south": "דרום",
    "judea-samaria": "יהודה ושומרון",
}

TAG_FIELDS = [
    "water",
    "viewpoint",
    "camping",
    "offroad",
    "strollerAccessible",
    "free",
]

TITLE_ALIASES: dict[str, str] = {
    "נחל שופט": "נחל השופט",
    "שמורת תל דן": "נחל דן",
    "המפל הנסתר פתח תקווה": "המפל הנסתר",
    "שמורת טבע נחל שורק": "שפך נחל שורק",
    "עין חרוד": "מעיין חרוד",
    "פארק נחל לכיש": "פארק לכיש",
    "מצפה נתן": "מצפה אלון",
    "נחל נקרות יקנעם": "פארק נחל קרת",
}

EXCEL_NAME_TO_SLUG: dict[str, str] = {
    "נחל שופט": "nahal-hashofet",
    "חוף מעגן מיכאל": "chvf-mgn-mykl",
}


def escape_ts(value: str) -> str:
    return value.replace("\\", "\\\\").replace('"', '\\"')


def normalize_title(value: str) -> str:
    text = value.strip()
    text = re.sub(r"[\s\-–—_/\\|]", "", text)
    text = re.sub(r"^ה", "", text)
    return text.casefold()


def parse_cell(value: object) -> str | bool | None:
    if value is None:
        return None
    raw = str(value).strip()
    if raw == "כן":
        return True
    if raw == "לא":
        return False
    if raw == "חלקי":
        return "partial"
    return raw or None


def load_site_trips() -> list[dict[str, str]]:
    trips: list[dict[str, str]] = []

    def extract(path: Path) -> None:
        if not path.is_file():
            return
        text = path.read_text(encoding="utf-8")
        pattern = re.compile(
            r'slug:\s*"([^"]+)"[\s\S]*?title:\s*"([^"]+)"[\s\S]*?region:\s*"([^"]+)"'
        )
        for match in pattern.finditer(text):
            trips.append(
                {
                    "slug": match.group(1),
                    "title": match.group(2),
                    "region": match.group(3),
                }
            )

    raw_chunk = TRIPS_FILE.read_text(encoding="utf-8")
    start = raw_chunk.index("const rawTrips")
    end = raw_chunk.index("export const trips")
    pattern = re.compile(
        r'slug:\s*"([^"]+)"[\s\S]*?title:\s*"([^"]+)"[\s\S]*?region:\s*"([^"]+)"'
    )
    for match in pattern.finditer(raw_chunk[start:end]):
        trips.append(
            {
                "slug": match.group(1),
                "title": match.group(2),
                "region": match.group(3),
            }
        )

    extract(VISITED_TRIPS_FILE)
    return trips


def build_title_index(trips: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    index: dict[str, dict[str, str]] = {}
    for trip in trips:
        keys = {
            normalize_title(trip["title"]),
            normalize_title(TITLE_ALIASES.get(trip["title"], trip["title"])),
        }
        for key in keys:
            if key:
                index.setdefault(key, trip)
    return index


def find_trip_for_excel_name(name: str, index: dict[str, dict[str, str]], trips: list[dict[str, str]]) -> dict[str, str] | None:
    slug_override = EXCEL_NAME_TO_SLUG.get(name)
    if slug_override:
        for trip in trips:
            if trip["slug"] == slug_override:
                return trip

    alias = TITLE_ALIASES.get(name, name)
    candidates = [name, alias]
    for candidate in candidates:
        hit = index.get(normalize_title(candidate))
        if hit:
            return hit

    normalized_name = normalize_title(name)
    for trip in trips:
        normalized_trip = normalize_title(trip["title"])
        if normalized_name == normalized_trip:
            return trip
        if normalized_name in normalized_trip or normalized_trip in normalized_name:
            return trip
    return None


def read_excel_rows(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        raise ValueError("Excel file is empty.")

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    missing = [column for column in COLUMNS if column not in headers]
    if missing:
        raise ValueError(f"Missing columns: {missing}. Found: {headers}")

    indices = {column: headers.index(column) for column in COLUMNS}
    entries: list[dict] = []

    for row in rows[1:]:
        if not row:
            continue
        name_raw = row[indices["שם המקום"]] if indices["שם המקום"] < len(row) else None
        if name_raw is None or not str(name_raw).strip():
            continue

        entry: dict = {"name": str(name_raw).strip()}
        for hebrew, field in COLUMNS.items():
            if field == "name":
                continue
            idx = indices[hebrew]
            entry[field] = parse_cell(row[idx] if idx < len(row) else None)

        region = str(entry.get("region") or "").strip()
        if region not in REGION_TO_SLUG:
            raise ValueError(f"Unknown region '{region}' for place '{entry['name']}'")

        entry["regionSlug"] = REGION_TO_SLUG[region]
        entry["regionTitle"] = region
        entries.append(entry)

    return entries


def render_sync_file(
    sync_by_slug: dict[str, dict],
    audit: dict,
) -> str:
    sync_lines = []
    for slug in sorted(sync_by_slug):
        entry = sync_by_slug[slug]
        tag_parts = []
        for field in TAG_FIELDS:
            value = entry["tags"].get(field)
            if value is True:
                tag_parts.append(f"{field}: true")
            elif value is False:
                tag_parts.append(f"{field}: false")
            elif value == "partial":
                tag_parts.append(f'{field}: "partial"')
        tags_body = ", ".join(tag_parts)
        sync_lines.append(
            f'  "{escape_ts(slug)}": {{ region: "{escape_ts(entry["regionTitle"])}", regionSlug: "{escape_ts(entry["regionSlug"])}", excelName: "{escape_ts(entry["excelName"])}", tags: {{ {tags_body} }} }},'
        )

    audit_json = json.dumps(audit, ensure_ascii=False, indent=2)

    return f"""/**
 * Trip region + filter tags synced from data/imports/places-filter.xlsx
 *
 * Generated by scripts/sync-places-filter.py — do not edit manually.
 */

import type {{ TripFilterTags }} from "@/lib/trip-filter-tags";

export type PlacesFilterSyncEntry = {{
  region: string;
  regionSlug: string;
  excelName: string;
  tags: TripFilterTags;
}};

export const placesFilterSyncBySlug: Record<string, PlacesFilterSyncEntry> = {{
{chr(10).join(sync_lines)}
}};

export const placesFilterAudit = {audit_json} as const;
"""


def main() -> int:
    parser = argparse.ArgumentParser(description="Sync places filter Excel into trip data")
    parser.add_argument("--input", type=Path, default=DEFAULT_EXCEL)
    args = parser.parse_args()

    excel_path = args.input
    if not excel_path.is_file():
        print(f"Excel file not found: {excel_path}", file=sys.stderr)
        return 1

    excel_rows = read_excel_rows(excel_path)
    site_trips = load_site_trips()
    title_index = build_title_index(site_trips)

    sync_by_slug: dict[str, dict] = {}
    matched_excel: list[str] = []
    unmatched_excel: list[str] = []
    matched_slugs: set[str] = set()

    for row in excel_rows:
        trip = find_trip_for_excel_name(row["name"], title_index, site_trips)
        if not trip:
            unmatched_excel.append(row["name"])
            continue

        matched_excel.append(row["name"])
        matched_slugs.add(trip["slug"])
        sync_by_slug[trip["slug"]] = {
            "regionTitle": row["regionTitle"],
            "regionSlug": row["regionSlug"],
            "excelName": row["name"],
            "tags": {field: row.get(field) for field in TAG_FIELDS},
        }

    unmatched_site = [
        trip["title"]
        for trip in site_trips
        if trip["slug"] not in matched_slugs
    ]

    region_counts = {slug: 0 for slug in REGION_TO_TITLE}
    for entry in sync_by_slug.values():
        region_counts[entry["regionSlug"]] = region_counts.get(entry["regionSlug"], 0) + 1

    filter_counts: dict[str, int] = {field: 0 for field in TAG_FIELDS}
    for entry in sync_by_slug.values():
        for field in TAG_FIELDS:
            value = entry["tags"].get(field)
            if value is True or value == "partial":
                filter_counts[field] += 1

    audit = {
        "excelTotal": len(excel_rows),
        "siteTotal": len(site_trips),
        "matched": len(matched_excel),
        "activeTrips": len(sync_by_slug),
        "hiddenUnmatchedSiteTrips": len(unmatched_site),
        "regionCounts": {
            "north": region_counts.get("north", 0),
            "hasharon": region_counts.get("hasharon", 0),
            "center": region_counts.get("center", 0),
            "jerusalem": region_counts.get("jerusalem", 0),
            "south": region_counts.get("south", 0),
            "judea-samaria": region_counts.get("judea-samaria", 0),
        },
        "filterCounts": filter_counts,
        "totalSynced": len(sync_by_slug),
        "missingFromExcel": unmatched_site,
        "missingFromSite": unmatched_excel,
    }

    OUTPUT_FILE.write_text(render_sync_file(sync_by_slug, audit), encoding="utf-8")

    print(f"Excel file: {excel_path.resolve()}")
    print(f"Wrote: {OUTPUT_FILE}")
    print("")
    print("=== AUDIT REPORT ===")
    print(f"Excel rows total: {audit['excelTotal']}")
    print(f"Active trips after sync: {audit['activeTrips']}")
    print(f"Hidden/unmatched site trips: {audit['hiddenUnmatchedSiteTrips']}")
    print(f"Excel rows missing matching site trip: {len(unmatched_excel)}")
    print("")
    print("Final region counts:")
    print(f"  North: {audit['regionCounts']['north']}")
    print(f"  Hasharon: {audit['regionCounts']['hasharon']}")
    print(f"  Center: {audit['regionCounts']['center']}")
    print(f"  Jerusalem: {audit['regionCounts']['jerusalem']}")
    print(f"  South: {audit['regionCounts']['south']}")
    if audit["regionCounts"]["judea-samaria"]:
        print(f"  Judea & Samaria: {audit['regionCounts']['judea-samaria']}")
    print("")
    print("Final filter counts (true + partial):")
    for field, count in audit["filterCounts"].items():
        print(f"  {field}: {count}")
    print("")
    if unmatched_site:
        print(f"Trips on site missing from Excel ({len(unmatched_site)}):")
        for title in unmatched_site:
            print(f"  - {title}")
    else:
        print("Trips on site missing from Excel: none")
    print("")
    if unmatched_excel:
        print(f"Trips in Excel missing from site ({len(unmatched_excel)}):")
        for title in unmatched_excel:
            print(f"  - {title}")
    else:
        print("Trips in Excel missing from site: none")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
